import os
from django.shortcuts import get_object_or_404
import openpyxl
import shutil 
import re
from openpyxl.worksheet.datavalidation import DataValidation
from main.models import Etudiant, Note, Semestre, Ue, Evaluation, Parcours, AnneeUniversitaire, Programme, Matiere
from django.db import DataError, IntegrityError, transaction
from datetime import datetime, timedelta

BASE_PATH = "media/excel_templates"

def trim_str(string):
    try:
        return float(string)
    except Exception as e:
        print(e)
        string = str(string).lower()
        string = re.sub(r'\s+', '', string.strip())
        string = re.sub(r'=', '', string)
        print(string)
        return string

def is_naturel(string):
    return string.isdecimal() and int(string) > 0

def convert_serial_temporel_number_to_date(numero_serie_temporelle):
    if str(numero_serie_temporelle).isdecimal():
        # Date de référence d'Excel
        date_reference = datetime(1899, 12, 30)  
        # Ajouter le nombre de jours au format de série temporelle à la date de référence
        date_resultat = date_reference + timedelta(days=numero_serie_temporelle)
        return date_resultat
    return numero_serie_temporelle

@transaction.atomic
def pre_load_ue_matiere_template_data_by_year(annees):
    folder_path = f"{BASE_PATH}/matieres_templates"
    os.mkdir(folder_path)
    template_path = f"{BASE_PATH}/ues_matieres_tmp.xlsx"
    
    for annee in annees:
        result_name = f"{folder_path}/ues_matieres_{annee}.xlsx"
        wb = openpyxl.load_workbook(filename=template_path)
        template_sheet = wb.active
        ues = Ue.objects.filter(programme__semestre__annee_universitaire__annee=annee.annee)
        new_sheets = []
        for ue in ues:
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = ue.codeUE
            new_sheet['A1'] = ue.libelle
            matieres = ue.matiere_set.all()
            if matieres:
                print(matieres)
                for count, matiere in enumerate(matieres):
                    new_sheet[f'A{count+3}'] = matiere.libelle
                    new_sheet[f'B{count+3}'] = matiere.coefficient
                    new_sheet[f'C{count+3}'] = matiere.minValue
                    new_sheet[f'D{count+3}'] = matiere.heures
                    new_sheet[f'E{count+3}'] = matiere.abbreviation
                    
            new_sheets.append(new_sheets)
        wb.save(result_name)   
        wb.close()

    shutil.make_archive(folder_path, 'zip', folder_path)
    
    for file_path in os.listdir(folder_path):
        os.remove(folder_path+"/"+file_path)
    os.rmdir(folder_path)
    
    return folder_path+".zip"

@transaction.atomic
def load_matieres_by_year(path, annee, batch_size=100):
    # Utiliser transaction.atomic pour améliorer les performances en effectuant toutes les opérations dans une seule transaction
    # Matiere.objects.filter(ue__programme__semestre__annee_universitaire=annee).delete()

    workbook = openpyxl.load_workbook(filename=path)

    # Utiliser prefetch_related pour minimiser les requêtes SQL lors de la récupération des objets Ue
    ues = Ue.objects.filter(codeUE__in=[sheet.title for sheet in workbook]).prefetch_related('matiere_set')
    for ue_sheet in workbook:
        try:
            ue = next(ue for ue in ues if ue.codeUE == ue_sheet.title)
        except StopIteration:
            continue

        for i, row in enumerate(ue_sheet.iter_rows(values_only=True, min_row=3, max_col=5)):
            libelle = row[0].strip()
            try:
                Matiere.objects.get_or_create(
                    libelle=libelle,
                    coefficient=int(trim_str(row[1])),
                    minValue=int(trim_str(row[2])),
                    heures=int(trim_str(row[3])),
                    abbreviation=trim_str(row[4]),
                    ue=ue,
                )
            except DataError as _:
                raise DataError("Vérifier les valeurs de vos champs !")
            except IntegrityError as _:
                raise IntegrityError("L'abbréviation d'une matière est unique !")

@transaction.atomic
def pre_load_note_ues_template_data(semestres):
    folder_path = f"{BASE_PATH}/notes_templates"
    template_path = f"{BASE_PATH}/ues_matieres_tmp.xlsx"
    
    if not os.path.exists(folder_path):
        os.mkdir(folder_path)
    
    template_folder_base_path ='media/excel_templates/notes_templates'
    
    for semestre in semestres: 
        try:
            programme = semestre.programme_set.all().get()
        except Exception:
            continue
        ues = programme.ues.all()  
        etudiants = semestre.etudiant_set.all()
        base_path =f'media/excel_templates/notes_templates/{semestre.libelle}'
        
        if not os.path.exists(base_path):
            os.mkdir(base_path)
        for ue in ues:
            print(ue)
            path = f'{base_path}/notes_{ue.codeUE}_{semestre.libelle}_{semestre.annee_universitaire}.xlsx'
            wb = openpyxl.load_workbook(filename=f"media/excel_templates/notes_ue_tmplt.xlsx")
            template_sheet = wb.active
            for matiere in ue.matiere_set.all():
                new_sheet = wb.copy_worksheet(template_sheet)
                new_sheet.title = matiere.codematiere
                new_sheet['B1'].value = ue.libelle
                new_sheet['B2'].value = semestre.annee_universitaire.annee
                new_sheet['B3'].value = semestre.libelle
                new_sheet['B4'].value = matiere.libelle
                min_row = 17
                max_row = min_row + len(etudiants)-1
                for i, row in enumerate(new_sheet.iter_rows(min_row=min_row, max_row=max_row, max_col=3)):
                    row[0].value = f"{etudiants[i].id} | {etudiants[i].nom}"
                    row[1].value = etudiants[i].prenom
                    row[2].value = ""
            wb.save(path)
            
    shutil.make_archive(template_folder_base_path, 'zip', template_folder_base_path)
    
    return template_folder_base_path+".zip"

@transaction.atomic
def load_notes_from_ue_file(path, ue, semestre):
    Evaluation.objects.filter(semestre=semestre).delete()
    
    evaluations = {
        "evaluation1" : [2, None],
        "evaluation2" : [3, None],
        "evaluation3" : [4, None],
        "evaluation4" : [5, None],
        "evaluation5" : [6, None],
        "evaluation6" : [7, None],
        "evaluation7" : [8, None],
        "rattrapage" : [9, None],
    }
    
    workbook = openpyxl.load_workbook(path)
    
    for sheet in workbook:
        matiere = sheet.title
        try:
            matiere = ue.matiere_set.get(codematiere=matiere)
        except Matiere.DoesNotExist:
            continue
        
        min_row = 7
        max_row = 13
        
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
            if row[3]:
                try:
                    evaluation_date = convert_serial_temporel_number_to_date(int(trim_str(row[1])))
                except ValueError as v:
                    raise v
                evaluation_name = str(row[2])
                evaluation_ponderation = int(trim_str(row[1]))
                cell_A_n = trim_str(row[0])
                rattrapage = cell_A_n == "rattrapage"
                evaluation, _ = Evaluation.objects.get_or_create(
                    libelle=evaluation_name, 
                    ponderation=evaluation_ponderation,
                    date=evaluation_date,
                    semestre=semestre,
                    matiere=matiere,
                    rattrapage=rattrapage,
                )
                evaluations[cell_A_n][1] = evaluation

        min_row = 17
        max_row = sheet.max_row

        for row in sheet.iter_rows(min_row=min_row, max_row=max_row ,values_only=True):
            matricule, _ = str(row[0]).split('|')

            try:
                matricule = trim_str(matricule).upper()
                etudiant = semestre.etudiant_set.get(id=matricule)
            except Exception as e:
                raise Etudiant.DoesNotExist(f'ID = {matricule}')
            for key in evaluations:
                data = evaluations[key]
                index = data[0]
                evaluation = data[1]
                valeur = trim_str(row[index])
                if evaluation and len(valeur) > 0:
                    valeur = int(valeur)
                    Note.objects.create(valeurNote=valeur, etudiant=etudiant, evaluation=evaluation)

@transaction.atomic
def pre_load_evaluation_template_data(matiere, semestre):
    path = 'media/excel_templates/evaluation_tmp.xlsx'
    
    wb = openpyxl.load_workbook(filename="media/excel_templates/evaluations.xlsx")
    ws = wb.active
    ws['B1'].value = semestre.annee_universitaire.annee
    ws['B2'].value = semestre.libelle
    ws['B3'].value = matiere.libelle
    ws['B5'].value = ""
    ws['B6'].value = ""
    ws['B7'].value = ""
    
    etudiants = matiere.get_etudiant_semestre(semestre).order_by("nom")
    
    min_row = 10
    max_row = min_row + len(etudiants)
    
    i = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row-1, max_col=3):
        row[0].value = etudiants[i].nom
        row[1].value = etudiants[i].prenom
        row[2].value = ""
        i += 1
    wb.save(path)
    
    return path

@transaction.atomic
def export_evaluation_data(matiere, semestre):
  
    evaluations=Evaluation.objects.filter(matiere=matiere,semestre=semestre)
    #print(evaluations)
   
    path = 'media/excel_templates/evaluations.xlsx'
    result_path=f'media/excel_templates/export/evaluations_{matiere.id}_{semestre}.xlsx'
    wb = openpyxl.load_workbook(filename=path)
    source=wb.active
    # source.title='evaluation_1'
    source['B1'].value = semestre.annee_universitaire.annee
    source['B2'].value = semestre.libelle
    source['B3'].value = matiere.libelle
    
    # evaluation_first=evaluations.first()
    # source['B4'].value = "Non" if evaluation_first.rattrapage else "Oui"

    # source['B5'].value =evaluation_first.date
    # source['B6'].value = evaluation_first.libelle
    # source['B7'].value = evaluation_first.ponderation
    
    #creation des feuilles
    for i in range(evaluations.count()):
        print("evalua: ",evaluations[i].libelle)
        new_sheet=wb.copy_worksheet(source)
        new_sheet.title=f'evaluation_{i+1}'
        evaluation=evaluations[i]
        new_sheet['B4'].value = "Non" if evaluation.rattrapage else "Oui"

        new_sheet['B5'].value =evaluation.date
        new_sheet['B6'].value = evaluation.libelle
        new_sheet['B7'].value = evaluation.ponderation
    #afficher les etudiants  
        etudiants = matiere.get_etudiant_semestre(semestre).order_by("nom")
      
        min_row = 10
        max_row = min_row + len(etudiants)
        
        i= 0
        for row in new_sheet.iter_rows(min_row=min_row, max_row=max_row-1, max_col=3):
            row[0].value = etudiants[i].nom
            row[1].value = etudiants[i].prenom
            
            #recupere la note de l'etudiant dans l'evaluation de cette matiere
            print(evaluation)
            note=evaluation.note_set.filter(etudiant=etudiants[i]).first()
            #notes=Note.objects.filter(etudiant=etudiants[i], evaluation__matiere=matiere,evaluation=evaluations[i])
            print(note)
            
            row[2].value =  note.valeurNote
            i+= 1
    
    
    
    wb.save(result_path)
    
    
    # wb = openpyxl.load_workbook(filename="media/excel_templates/evaluations.xlsx")
    # ws = wb.active
    # ws['B1'].value = semestre.annee_universitaire.annee
    # ws['B2'].value = semestre.libelle
    # ws['B3'].value = matiere.libelle
    
    
    # for index, sublist in enumerate(table):
    # # Créer une nouvelle feuille de calcul avec un nom unique
    #     sheet_name = f'Sheet{index + 1}'
    #     sheet = wb.create_sheet(title=sheet_name)
    
    # wb = openpyxl.load_workbook(filename="media/excel_templates/evaluations.xlsx")
    # ws = wb.active
    # ws['B1'].value = semestre.annee_universitaire.annee
    # ws['B2'].value = semestre.libelle
    # ws['B3'].value = matiere.libelle
    # ws['B5'].value = ""
    # ws['B6'].value = ""
    # ws['B7'].value = ""
    
    # etudiants = matiere.get_etudiant_semestre(semestre).order_by("nom")
    
    # min_row = 10
    # max_row = min_row + len(etudiants)
    
    # i = 0
    # for row in ws.iter_rows(min_row=min_row, max_row=max_row-1, max_col=3):
    #     row[0].value = etudiants[i].nom
    #     row[1].value = etudiants[i].prenom
    #     row[2].value = ""
    #     i += 1
    
    
    return result_path


@transaction.atomic
def load_notes_from_evaluation(path, matiere=None, semestre=None):
    # Charger le fichier excel des différentes notes d'une matière
    wb = openpyxl.load_workbook(path)
    wb.iso_dates = True

    # Parcourir les evaluations
    for sheet in wb:
        columns = {
            'nom_cell' : 0,
            'prenom_cell' : 1,
            'evaluations' : [],
        }
        
        ws = sheet
        if sheet.title != "evaluation" :
            print("sheet: ",sheet)
            annee = str(ws['B1'].value)
            matiere = None
            semestre = None
            # if not semestre:
            #     semestre = str(ws['B2'].value)
            #     matiere = Matiere.objects.get(libelle=matiere)
            # if not matiere:
            #     annee = AnneeUniversitaire.objects.get(annee=annee)
            #     matiere = str(ws['B3'].value)
            #     semestre = Semestre.objects.get(libelle=semestre, annee_universitaire__id=annee.id)
            
            annee = AnneeUniversitaire.objects.get(annee=annee)
            semestre = str(ws['B2'].value)
            matiere = str(ws['B3'].value)
            matiere = Matiere.objects.get(libelle=matiere)
            semestre = Semestre.objects.get(libelle=semestre, annee_universitaire__id=annee.id)
            
        
            rattrapage = False
            print(ws['B5'].value)
            evaluation_date = convert_serial_temporel_number_to_date(ws['B5'].value)
            evaluation_name = str(ws['B6'].value)
            evaluation_ponderation = float(str(ws['B7'].value))
            
            #evaluation_date = evaluation_date.date()
            
            if (matiere.ponderation_restante(semestre) - evaluation_ponderation) >= 0:
                evaluation, _ = Evaluation.objects.get_or_create(
                        libelle=evaluation_name, 
                        ponderation=evaluation_ponderation,
                        date=evaluation_date,
                        semestre=semestre,
                        matiere=matiere,
                        rattrapage=rattrapage,
                    )
                
                etudiants = matiere.get_etudiant_semestre(semestre)
                print(etudiants)
                print(matiere, " ", semestre)
                min_row = 10
                max_row = min_row + len(etudiants)
                
                for row in sheet.iter_rows(min_row=min_row, max_row=max_row-1, max_col=3, values_only=True):
                    nom = row[0]
                    prenom = row[1]
                    if nom==prenom==None:
                        continue
                    username = (nom + prenom).lower()
                    username = username.replace(" ", "")
                    print(username)
                    etudiant = etudiants.get(user__username=username)
                    if Note.objects.filter(etudiant=etudiant, evaluation=evaluation):
                        continue
                    Note.objects.create(valeurNote=int(trim_str(row[2])), etudiant=etudiant, evaluation=evaluation)
            else:
                raise ValueError(f"Erreur de pondération pour l'évaluation : {evaluation_name} ")
# mangbayét-nambénédicta
# mangbayet-nambénédicta
@transaction.atomic
def pre_load_maquette(annees):
    path = 'media/excel_templates/maquette_general_[annee].xlsx'
    folder_path = f"{BASE_PATH}/maquette_templates"
    os.mkdir(folder_path)
    template_path = f"{BASE_PATH}/ues_matieres_tmp.xlsx"
    
    for annee in annees:
        result_path = f"{folder_path}/maquette_general_{annee}.xlsx"
        
        wb = openpyxl.load_workbook(path)
        semestres = annee.semestre_set.all().prefetch_related('programme_set')
        for sheet in wb:
            semestre = sheet.title
            semestre = semestres.filter(libelle=semestre).first()
            programmes = semestre.programme_set.all()
            
            valeurs_possibles_type = [ value[1] for value in Ue.TYPES ]
            valiation_type = DataValidation(type="list", formula1='"'+ ','.join(map(str, valeurs_possibles_type))+'"', allow_blank=True)
            valeurs_possibles_niveau = [ value[1] for value in Ue.TYPES_NIVEAU ]
            valiation_niveau = DataValidation(type="list", formula1='"'+ ','.join(map(str, valeurs_possibles_niveau))+'"', allow_blank=True)
            
            sheet.add_data_validation(valiation_type)
            sheet.add_data_validation(valiation_niveau)
            count = 3
            if programmes:
                programme = programmes.first()
                for ue  in programme.ues.all():
                    cell_a = sheet[f'A{count}']
                    cell_a.value = ue.libelle
                    cell_b = sheet[f'B{count}']
                    cell_b.value = valeurs_possibles_type[0]
                    valiation_type.add(cell_b)
                    
                    cell_c = sheet[f'C{count}']
                    cell_c.value = valeurs_possibles_niveau[0]
                    valiation_niveau.add(cell_c)
                    
                    cell_d = sheet[f'D{count}']
                    cell_d.value = ue.nbreCredits
                    cell_e = sheet[f'E{count}']
                    cell_e.value = ue.heures
                    count += 1
                    
            start = count
            end = 21
            for i in range(start, end):
                cell_a = sheet[f'A{i}']
                cell_a.value = ""
                cell_b = sheet[f'B{i}']
                cell_b.value = valeurs_possibles_type[0]
                valiation_type.add(cell_b)
                
                cell_c = sheet[f'C{i}']
                cell_c.value = valeurs_possibles_niveau[0]
                valiation_niveau.add(cell_c)
                
            wb.save(result_path)
        
    shutil.make_archive(folder_path, 'zip', folder_path)
    
    for file_path in os.listdir(folder_path):
        os.remove(folder_path+"/"+file_path)
    os.rmdir(folder_path)
    
    return folder_path+".zip"

@transaction.atomic         
def load_maquette(path, annee):
    try:
        workbook = openpyxl.load_workbook(filename=path, read_only=True)

        semestre_ue = {}
        for sheet in workbook:
            semestre = sheet.title.lower()
            semestre_ue[semestre] = []
            print(sheet.max_row)
            for i in range(3, sheet.max_row):
                libelle = sheet[f'A{i}'].value
                type = sheet[f'B{i}'].value 
                niveau = sheet[f'C{i}'].value
                nbreCredits = sheet[f'D{i}'].value
                heures = sheet[f'E{i}'].value
                if libelle and type and niveau and nbreCredits and heures:
                    libelle = libelle.strip()
                    nbreCredits = trim_str(nbreCredits)
                    heures = trim_str(heures)
                    niveau = trim_str(niveau)
                    #print("libelle", libelle, " type" , type, " niveau", niveau, " nbreCredits", nbreCredits, " heures", heures)
                    if is_naturel(nbreCredits) and is_naturel(heures):
                        ue, _ = Ue.objects.get_or_create(libelle=libelle, type=type, niveau=niveau, nbreCredits=nbreCredits, heures=heures)
                        semestre_ue[semestre].append(ue)
                    else:
                        raise ValueError("Vous avez des valeurs négatives dans votre fichier")
                
        parcours = Parcours.objects.all().first()
        semestres = annee.semestre_set.all()

        for semestre in semestres:
            programme, _ = Programme.objects.get_or_create(
                semestre=semestre, parcours=parcours)
            programme.ues.set(semestre_ue[semestre.libelle.lower()])
    except ValueError as e:
        raise e
    except Exception as e:
        print(e)
        raise ValueError("Le fichier n'est pas dans le bon format")
    finally:
        if "workbook" in locals():
            workbook.close()
            
@transaction.atomic
def load_notes_from_matiere(path):
    Evaluation.objects.all().delete()
    # Charger le fichier excel des différentes notes d'une matière
    workbook = openpyxl.load_workbook(path)
    
    # Initialiser les variable de base
    annee = None
    semestre = None
    matiere = None
    
    for sheet in workbook:
        columns = {
            'prenom_cell' : 0,
            'nom_cell' : 1,
            'evaluations' : [],
        }
        nb_evaluation = columns['nom_cell']
        for row in sheet.iter_rows(values_only=True):
            if row[0]:
                firts_row_elt_value = trim_str(row[0])
                second_row_elt_value = row[1]
                if "annee:" in firts_row_elt_value:
                    annee = trim_str(second_row_elt_value)
                    annee = AnneeUniversitaire.objects.get(annee=annee)
                elif "semestre:" in firts_row_elt_value:
                    if annee:
                        semestre = trim_str(second_row_elt_value)
                        print(semestre)
                        semestre = annee.semestre_set.get(libelle=semestre.upper())
                elif "matière:" in firts_row_elt_value:
                    matiere = second_row_elt_value.strip()
                    matiere = Matiere.objects.get(libelle=matiere, ue__programme__semestre=semestre)
                elif 'évaluation' in firts_row_elt_value:
                    # Créer les evaluations 
                    if annee and semestre and matiere:
                        print(row)
                        evaluation = Evaluation.objects.create(libelle=trim_str(row[2]), matiere=matiere, semestre=semestre, ponderation=int(trim_str(row[3])), date="2023-11-12")
                        nb_evaluation += 1
                        columns['evaluations'].append({ 
                                                        'evaluation' : evaluation,
                                                        'cell' : nb_evaluation,
                                                        })
                    else:
                        raise Exception("Le format de votre fichier est ivalide !")
                elif firts_row_elt_value != "prénom":
                        nom = str(row[columns['nom_cell']])
                        prenom = str(row[columns['prenom_cell']])
                        try:
                            etudiant = Etudiant.objects.get(nom__icontains=nom, prenom__icontains=prenom)
                            print(etudiant)
                            for evaluataion_data in columns['evaluations']:
                                Note.objects.create(valeurNote=int(trim_str(row[evaluataion_data['cell']])), etudiant=etudiant, evaluation=evaluataion_data['evaluation'])
                        except Exception as e:
                            raise Etudiant.DoesNotExist(f'username = {firts_row_elt_value}')

def run():
    #clean_data_base()
    print("::::: Import begining :::::")
    semestre = Semestre.objects.get(annee_universitaire__annee=2023, libelle="S1")
    # print(semestre)
    # matiere = Matiere.objects.get()
    # Note.objects.filter(valeurNote=float("-1"))
    # for note in notes:
    #     print(note)
    print("Hello")


